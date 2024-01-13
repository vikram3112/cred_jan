import openpyxl


def rows_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def ReadData(file, sheetname, rownno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownno, column=colno).value


def WriteData(file, sheetname, rownno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownno, column=colno).value = data
    workbook.save(file)
